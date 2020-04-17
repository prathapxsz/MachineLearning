#Importing required packages
import requests
from bs4 import BeautifulSoup
import openpyxl as xl
import time
#url set to url variable from where we gonna scrape data
url = 'https://finance.yahoo.com/quote/AAPL?p=AAPL'
while 4>3:
    
     #requesting data from url

    data = requests.get(url)
    if data.status_code == 200:
    #Getting html data and reading it using lxml as parser
        soup = BeautifulSoup(data.text, 'lxml')

        #Getting first table from soup data
        table = soup.find('table', class_='W(100%)')
        tr = table.findAll('td', class_='C($primaryColor) W(51%)')
        #Iterating through table data to get keys and storing them to a new list called keys
        keys = []
        for row in tr:
            keys.append(row.text)
        #Iterating through table data again to get values for above keys and storing them to a new list called values
        tr2 = table.findAll('td', class_='Ta(end) Fw(600) Lh(14px)')
        values = []
        for row in tr2:
            values.append(row.text)

        #Getting second table from soup data
        second = soup.find('table', class_='W(100%) M(0) Bdcl(c)')
        td = second.findAll('td', class_='C($primaryColor) W(51%)')
        #Iterating through data to get keys and storing them to keys list
        for row2 in td:
            keys.append(row2.text)
        #iterating through data to get values of above keys and storing them to values list
        td2 = second.findAll('td', class_='Ta(end) Fw(600) Lh(14px)')
        for row3 in td2:
            values.append(row3.text)
    
        keys_store = keys
        values_store = values
        keys_store.append("url")
        values_store.append(url)
        #Above keys and value are stored into dictionary and printed 
        '''
        my_dict = {}
        my_dict = zip(keys, values)
        for keys, values in my_dict:
            print("%s : %s"%(keys,values))
        '''
        #Creating a excel workbook as test.xlsv
        wb = xl.Workbook()
        wb.save('./test.xlsx')
        #Loading the workbook and creating a new sheet as Stock Data
        wb = xl.load_workbook('./test.xlsx')
        main_sheet = wb.create_sheet()
        main_sheet.title = 'Stock Data'
        #Iterating through keys and values and storing them into excel
        key_row = 1
        for key in keys_store:
            main_sheet.cell(column=1, row=key_row, value=key)
            key_row += 1
        value_row = 1
        for value in values_store:
            main_sheet.cell(column=5, row=value_row, value=value)
            value_row += 1
        wb.save('test.xlsx')
        print("Updated")
        time.sleep(60)
            
    else:
        print("Failed to get page. Trying again")
            
